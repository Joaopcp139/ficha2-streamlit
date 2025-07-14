import streamlit as st
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from io import BytesIO
import re
from datetime import datetime

st.set_page_config(page_title="Gerador de Ficha Técnica", layout="centered")
st.title("📄 Gerador de Ficha Técnica a partir da OP")

# Upload da OP
uploaded_pdf = st.file_uploader("📎 Envie a OP em PDF", type="pdf")

# Tipo de ficha
ficha_tipo = st.radio("Tipo de ficha técnica:", ["SACO", "FILME"])

# Botão para processar
if uploaded_pdf and st.button("🔄 Gerar ficha técnica"):
    try:
        # 1. Leitura do PDF
        reader = PdfReader(uploaded_pdf)
        texto = ""
        for page in reader.pages:
            texto += page.extract_text() + "\n"

        # 2. Extração dos campos
        def extrair(padrao, texto, default=""):
            match = re.search(padrao, texto)
            return match.group(1).strip() if match else default

        dados = {
            "cliente": extrair(r"Cliente:\s*(.+)", texto),
            "produto": extrair(r"Produto:\s*(.+)", texto),
            "codigo_produto": extrair(r"(\d{5,})\s*-\s*", texto),
            "data_pedido": extrair(r"Data do Pedido:\s*(\d{2}/\d{2}/\d{4})", texto),
            "data_entrega": extrair(r"Data de Entrega:\s*(\d{2}/\d{2}/\d{4})", texto),
            "pedido_numero": extrair(r"Pedido Nº:\s*(\d+)", texto),
            "largura": extrair(r"Largura:\s*(\d+)", texto),
            "espessura": extrair(r"Espessura:\s*([0-9,\.]+)", texto),
            "passo": extrair(r"Passo:\s*(\d+)", texto),
            "cilindro": extrair(r"Cilindro:\s*(\d+)", texto),
            "quantidade_kg": extrair(r"Quantidade \(KG\):\s*([0-9\.]+)", texto),
            "quantidade_bobinas": extrair(r"Quantidade de bobinas:\s*(\d+)", texto),
            "tubete": "Yes" if "Tubete 3: Sim" in texto else "No",
            "laminado": "Sim" if "Laminado: Sim" in texto else "Não",
            "sanfona": "Sim" if "Sanfona Sim" in texto else "Não",
            "materia_prima": "Yes" if "Matéria-prima PE: Sim" in texto else "No",
            "frente1": "Yes" if "Frente 1: Yes" in texto else "No",
            "oc": extrair(r"OC:\s*(\d+)", texto),
        }

        # 3. Carrega modelo correto
        modelo_path = "FILME.xlsx" if ficha_tipo == "FILME" else "SACO.xlsx"
        wb = load_workbook(modelo_path)
        ws = wb.active

        # 4. Preenchimento (com fallback para células vazias)
        campos = {
            "D6": dados["cliente"],
            "F6": dados["codigo_produto"],
            "D7": dados["produto"],
            "B13": dados["largura"],
            "D13": dados["passo"],
            "F13": dados["espessura"],
            # Adicione mais mapeamentos conforme necessário
        }
        
        # Insere data atual na célula L2
        hoje = datetime.now().strftime("%d/%m/%Y")
        ws["L2"] = hoje

        # Preenche células apenas se houver valor
        for celula, valor in campos.items():
            if valor:  # Só preenche se não for vazio
                ws[celula] = valor

        # 5. Exporta para download
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("✅ Ficha técnica gerada com sucesso!")
        st.download_button(
            label="📥 Baixar ficha preenchida",
            data=output,
            file_name="ficha_tecnica_preenchida.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Erro ao gerar ficha técnica: {e}")
